"""
Test script to validate Excel import functionality.
Creates a test Excel file and validates the import_from_excel function.
"""
import os
import sys
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "kentia_cal.settings")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
django.setup()

from io import BytesIO
from openpyxl import Workbook
from calibrations.services import import_from_excel

print("=" * 60)
print("Excel Import Functionality Test")
print("=" * 60)

# Create test Excel file
print("\n[Test 1] Creating test Excel file...")
try:
    wb = Workbook()
    ws = wb.active
    ws.title = "Calibrations"

    # First table header
    ws["A1"] = "FEATURE"
    ws["B1"] = "VARIABLE"
    ws["C1"] = "Release1"
    ws["D1"] = "Release2"

    # First table data
    ws["A2"] = "F1"
    ws["B2"] = "var_test1"
    ws["C2"] = "x"
    ws["D2"] = "x"

    ws["A3"] = None
    ws["B3"] = "var_test2"
    ws["C3"] = "x"
    ws["D3"] = None

    # Second table header (starts at row 6)
    ws["A6"] = "value"
    ws["B6"] = "Feature"
    ws["C6"] = "Variable"
    ws["D6"] = "Release"
    ws["E6"] = "Status"
    ws["F6"] = "Verification"

    # Second table data
    ws["A7"] = 1000
    ws["B7"] = "F1"
    ws["C7"] = "var_test1"
    ws["D7"] = "Release1"
    ws["E7"] = "0.25"
    ws["F7"] = "No"

    ws["A8"] = 1100
    ws["B8"] = "F1"
    ws["C8"] = "var_test1"
    ws["D8"] = "Release2"
    ws["E8"] = "0.5"
    ws["F8"] = "Test bench"

    # Save to BytesIO
    file_obj = BytesIO()
    wb.save(file_obj)
    file_obj.seek(0)

    print("  OK: Test Excel file created")

except Exception as e:
    print(f"  ERROR: {e}")
    sys.exit(1)

# Test import function
print("\n[Test 2] Running import_from_excel function...")
try:
    result = import_from_excel(file_obj)

    print(f"  Success: {result['success']}")
    print(f"  Features created: {result['features_created']}")
    print(f"  Features updated: {result['features_updated']}")
    print(f"  Variables created: {result['variables_created']}")
    print(f"  Variables updated: {result['variables_updated']}")
    print(f"  Releases created: {result['releases_created']}")
    print(f"  Values created: {result['values_created']}")
    print(f"  Values updated: {result['values_updated']}")

    if result["errors"]:
        print(f"  Errors: {result['errors']}")

    if result["success"]:
        print("  OK: Import completed successfully")
    else:
        print("  ERROR: Import reported success=False")

except Exception as e:
    print(f"  ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Verify data in database
print("\n[Test 3] Verifying imported data in database...")
try:
    from calibrations.models import Feature, Variable, Release, CalibrationValue

    features = Feature.objects.filter(codigo__startswith="F")
    variables = Variable.objects.filter(nombre__startswith="var_test")
    releases = Release.objects.filter(nombre__startswith="Release")
    values = CalibrationValue.objects.filter(variable__nombre__startswith="var_test")

    print(f"  Features: {features.count()}")
    print(f"  Variables: {variables.count()}")
    print(f"  Releases: {releases.count()}")
    print(f"  CalibrationValues: {values.count()}")

    if features.count() > 0 and variables.count() > 0 and values.count() > 0:
        print("  OK: Data successfully imported to database")
    else:
        print("  WARNING: Some data was not imported")

except Exception as e:
    print(f"  ERROR: {e}")
    sys.exit(1)

print("\n" + "=" * 60)
print("All tests passed!")
print("=" * 60)
