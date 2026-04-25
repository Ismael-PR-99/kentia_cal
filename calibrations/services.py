import logging
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from django.core.exceptions import MultipleObjectsReturned

from .models import CalibrationValue, Feature, Release, Variable

logger = logging.getLogger(__name__)


def import_from_excel(file_obj) -> dict:
    """
    Parse Excel file and import calibration data.
    Expected format:
    - First table: Feature (A), Variable (B), Releases (C+) with 'x' marker
    - Second table: value, status, verification, unit, description, release, responsible

    Returns dict with import summary:
    {
        "success": bool,
        "features_created": int,
        "features_updated": int,
        "variables_created": int,
        "variables_updated": int,
        "releases_created": int,
        "values_created": int,
        "values_updated": int,
        "errors": [str],
    }
    """
    summary = {
        "success": False,
        "features_created": 0,
        "features_updated": 0,
        "variables_created": 0,
        "variables_updated": 0,
        "releases_created": 0,
        "values_created": 0,
        "values_updated": 0,
        "errors": [],
    }

    try:
        # Load workbook
        file_bytes = file_obj.read()
        workbook = load_workbook(BytesIO(file_bytes))
        sheet = workbook.active

        # Parse first table: features and variable-release assignments
        features_map = {}  # {feature_code: Feature}
        variables_map = {}  # {(feature_code, var_name): Variable}
        releases_in_table = []  # [Release names from headers]
        variable_release_map = {}  # {(var_id, release_name): marked}

        # Find header row (row with releases)
        header_row = None
        for row_idx in range(1, sheet.max_row + 1):
            cell_a = sheet[f"A{row_idx}"].value
            if cell_a and str(cell_a).upper() == "FEATURE":
                header_row = row_idx
                break

        if not header_row:
            summary["errors"].append("Could not find header row with 'FEATURE'")
            return summary

        # Extract release names from header
        for col_idx in range(3, 20):  # Columns C onwards
            cell_val = sheet.cell(header_row, col_idx).value
            if cell_val:
                releases_in_table.append(str(cell_val).strip())

        # Create/update releases
        for release_name in releases_in_table:
            release, created = Release.objects.update_or_create(
                nombre=release_name,
                defaults={"fecha": date.today(), "descripcion": "Imported from Excel"},
            )
            if created:
                summary["releases_created"] += 1

        # Parse data rows (skip header)
        current_feature = None
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            cell_a = sheet[f"A{row_idx}"].value
            cell_b = sheet[f"B{row_idx}"].value
            cell_c = sheet[f"C{row_idx}"].value

            # End of first table if all cells empty
            if not cell_a and not cell_b and not cell_c:
                break

            # Feature in column A (may span multiple rows)
            if cell_a:
                feature_code = str(cell_a).strip()
                feature, created = Feature.objects.update_or_create(
                    codigo=feature_code,
                    defaults={"nombre": f"Feature {feature_code}", "descripcion": "Imported from Excel"},
                )
                if created:
                    summary["features_created"] += 1
                else:
                    summary["features_updated"] += 1
                current_feature = feature
                features_map[feature_code] = feature

            # Variable in column B
            if cell_b and current_feature:
                var_name = str(cell_b).strip()
                variable, created = Variable.objects.update_or_create(
                    feature=current_feature,
                    nombre=var_name,
                    defaults={
                        "descripcion": "Imported from Excel",
                        "unidad": "",
                        "responsable": Variable.Responsible.SUPPLIER,
                        "dimension_type": Variable.DimensionType.SCALAR_1X1,
                    },
                )
                if created:
                    summary["variables_created"] += 1
                else:
                    summary["variables_updated"] += 1
                variables_map[(current_feature.codigo, var_name)] = variable

                # Mark which releases this variable belongs to
                for col_idx, release_name in enumerate(releases_in_table, start=3):
                    cell_val = sheet.cell(row_idx, col_idx).value
                    if cell_val and str(cell_val).strip().lower() == "x":
                        variable_release_map[(variable.id, release_name)] = True

        # Parse second table: calibration values
        # Find start of second table (look for "value" or similar header)
        second_table_row = None
        for row_idx in range(header_row + 20, sheet.max_row + 1):
            cell_a = sheet[f"A{row_idx}"].value
            if cell_a and str(cell_a).lower() in ["value", "valor"]:
                second_table_row = row_idx
                break

        if second_table_row:
            second_table_headers = {}
            for col_idx in range(1, sheet.max_column + 1):
                header_value = sheet.cell(second_table_row, col_idx).value
                if header_value:
                    second_table_headers[str(header_value).strip().lower()] = col_idx

            value_col = second_table_headers.get("value") or second_table_headers.get("valor")
            status_col = second_table_headers.get("status (maturity)") or second_table_headers.get("status")
            verification_col = second_table_headers.get("verification") or second_table_headers.get("verificacion")
            unit_col = second_table_headers.get("unit") or second_table_headers.get("unidad")
            description_col = second_table_headers.get("description") or second_table_headers.get("descripcion")
            release_col = second_table_headers.get("release")
            responsible_col = second_table_headers.get("responsible") or second_table_headers.get("responsable")
            variable_col = second_table_headers.get("variable")

            if not value_col or not release_col or not variable_col:
                summary["errors"].append("La cabecera de la segunda tabla no contiene las columnas mínimas requeridas.")
                summary["success"] = False
                return summary

            # Parse calibration values
            for row_idx in range(second_table_row + 1, sheet.max_row + 1):
                cell_value = sheet.cell(row_idx, value_col).value if value_col else None
                cell_status = sheet.cell(row_idx, status_col).value if status_col else None
                cell_verification = sheet.cell(row_idx, verification_col).value if verification_col else None
                cell_unit = sheet.cell(row_idx, unit_col).value if unit_col else None
                cell_description = sheet.cell(row_idx, description_col).value if description_col else None
                cell_release = sheet.cell(row_idx, release_col).value if release_col else None
                cell_responsible = sheet.cell(row_idx, responsible_col).value if responsible_col else None
                cell_variable = sheet.cell(row_idx, variable_col).value if variable_col else None

                # End of table
                if not any([cell_value, cell_variable, cell_release, cell_status, cell_verification]):
                    break

                if cell_variable and cell_release and cell_value is not None:
                    var_name = str(cell_variable).strip()
                    release_name = str(cell_release).strip()
                    valor = float(cell_value) if cell_value else 0
                    status = str(cell_status).strip() if cell_status else "0.25"
                    verification = str(cell_verification).strip() if cell_verification else "No"
                    unidad = str(cell_unit).strip() if cell_unit else ""
                    descripcion = str(cell_description).strip() if cell_description else ""
                    responsable = str(cell_responsible).strip() if cell_responsible else ""

                    try:
                        variable = Variable.objects.get(nombre=var_name)
                    except MultipleObjectsReturned:
                        variable = Variable.objects.filter(nombre=var_name).order_by("id").first()
                    except Variable.DoesNotExist:
                        variable = None

                    try:
                        release = Release.objects.get(nombre=release_name)

                        if not variable:
                            raise Variable.DoesNotExist()

                        variable_updated = False
                        if unidad and variable.unidad != unidad:
                            variable.unidad = unidad
                            variable_updated = True
                        if descripcion and variable.descripcion != descripcion:
                            variable.descripcion = descripcion
                            variable_updated = True
                        if responsable:
                            valid_responsibilities = {choice for choice, _label in Variable.Responsible.choices}
                            normalized_responsable = responsable
                            if responsable.upper() == "SUPPLIER":
                                normalized_responsable = Variable.Responsible.SUPPLIER
                            elif responsable.upper() == "OEM":
                                normalized_responsable = Variable.Responsible.OEM
                            if normalized_responsable in valid_responsibilities and variable.responsable != normalized_responsable:
                                variable.responsable = normalized_responsable
                                variable_updated = True

                        if variable_updated:
                            variable.save(update_fields=["unidad", "descripcion", "responsable", "updated_at"])

                        _, created = CalibrationValue.objects.update_or_create(
                            variable=variable,
                            release=release,
                            defaults={
                                "valor": valor,
                                "status_madurez": status,
                                "verificacion": verification,
                                "notas": "Imported from Excel",
                            },
                        )
                        if created:
                            summary["values_created"] += 1
                        else:
                            summary["values_updated"] += 1
                    except (Variable.DoesNotExist, Release.DoesNotExist):
                        summary["errors"].append(
                            f"Row {row_idx}: Variable ({var_name}) or Release ({release_name}) not found"
                        )

        summary["success"] = True
        return summary

    except Exception as e:
        logger.error(f"Excel import error: {str(e)}")
        summary["errors"].append(f"Import failed: {str(e)}")
        return summary
